import pdfreader
import re
import os
import openpyxl

class Result:
    def __init__(self,exBtw,btw,total,date):
        self.exBtw = exBtw
        self.btw = btw
        self.total = total
        self.date = date
    def toList(self):
        return [self.date,"","",self.exBtw,self.btw,self.total] 

#returns a result object form a given pdf
def analysePDF(pathName):
    print(f"processing: {pathName}")
    content = getPdfText(pathName)
    #indexed relative to the end of the list, because the index compared to the start of the document is not constant
    totalEur = content[-17]
    exBtw = content[-5]
    btw = content[-4]
    invoiceDate = content[7]
    #following asserts are to verify input is in correct format by asserting that al instances of a number are equal 
    assert totalEur == content[-20],"totals dont match"
    assert exBtw == content[-8] , "exBtw values dont match"
    assert btw == content[-7], "btw values dont match"
    invoiceDateRe = re.compile("Invoice Date : \d{4}-\d{2}-\d{2}")
    assert invoiceDateRe.match(invoiceDate), "invoicedate not in right format"
    #put the date in the DD-MM-YYYY format instead of the YYYY-DD-MM format that aliexpres uses
    date = invoiceDate[20:25]+"-"+invoiceDate[15:19]
    newResult = Result(exBtw,btw,totalEur,date)
    return newResult

#get an array of strings of a given pdf file
def getPdfText(pathName):
    res = []
    fd = open(pathName,"rb")
    doc = pdfreader.PDFDocument(fd)
    viewer = pdfreader.SimplePDFViewer(fd)
    page_length = len([p for p in doc.pages()])
    for i in range(page_length):
        viewer.navigate(i+1)
        viewer.render()
        viewer.canvas.strings
        pageContent = viewer.canvas.strings
        res += pageContent
    return res
        
#get results from all pdf files in a directory
def getResults(directory):
    bestandjes = getFilesInDirectory(directory)
    validBestandjeRE = re.compile("\d+\_payment.pdf")
    results = []
    for bestandje in bestandjes:
        assert validBestandjeRE.match(bestandje), "bestand naam niet in correct formaat"
        res = analysePDF(os.path.join(directory,bestandje))
        results.append(res)
    return results

#returns a list of all the files in a directory
def getFilesInDirectory(directory):
    directoryContent = os.listdir(directory)
    files = [f for f in directoryContent if os.path.isfile(os.path.join(directory,f))]
    return files

#puts a list of results in to an openpyxl workbook
def resultatenToWorkbook(resultaten):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Datum","","","exBtw","Btw","Totaal"])
    for res in resultaten:
        worksheet.append(res.toList())
    return workbook
    
def main():
    directory = input("name of folder containing payment pdfs: ")
    output = input("name of output excel file (should end with .xlsx): ")
    resultaten = getResults(directory)
    workbook = resultatenToWorkbook(resultaten)
    workbook.save(output)
    print(f"saved at: {os.path.abspath(output)}")

if __name__=="__main__":
    main()
